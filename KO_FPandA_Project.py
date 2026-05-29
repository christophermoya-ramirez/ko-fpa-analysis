# Coca-Cola (KO) FP&A Analysis Tool
# Author: Christopher Moya Ramirez
# Description: Automated financial analysis tool that pulls
# live KO data, calculates key FP&A metrics, compares against
# PEP, and exports a polished Excel report.

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# STEP 1 — Pull data for both companies

print("Pulling live data from Yahoo Finance...")

ko = yf.Ticker("KO")
pep = yf.Ticker("PEP")

print("KO data loaded.")
print("PEP data loaded.")
print("Building report...\n")

# STEP 2 — Pull and clean income statement data

# Coca-Cola income statement
ko_income = ko.financials.T
ko_income.index = ko_income.index.year
ko_metrics = ko_income[["Total Revenue", "Gross Profit", "Operating Income", "Net Income"]].copy()
ko_metrics = (ko_metrics / 1_000_000).round(1)

# PepsiCo income statement
pep_income = pep.financials.T
pep_income.index = pep_income.index.year
pep_metrics = pep_income[["Total Revenue", "Gross Profit", "Operating Income", "Net Income"]].copy()
pep_metrics = (pep_metrics / 1_000_000).round(1)

# STEP 3 — Calculate margins for both companies

for metrics in [ko_metrics, pep_metrics]:
    metrics["Gross Margin %"] = ((metrics["Gross Profit"] / metrics["Total Revenue"]) * 100).round(1)
    metrics["Operating Margin %"] = ((metrics["Operating Income"] / metrics["Total Revenue"]) * 100).round(1)
    metrics["Net Margin %"] = ((metrics["Net Income"] / metrics["Total Revenue"]) * 100).round(1)
    metrics["Revenue Growth %"] = (metrics["Total Revenue"].pct_change(-1) * 100).round(1)

print("=== KO Income Statement ($M) ===")
print(ko_metrics[["Total Revenue", "Gross Profit", "Operating Income", "Net Income"]].dropna().to_string())

print("\n=== KO Margins ===")
print(ko_metrics[["Gross Margin %", "Operating Margin %", "Net Margin %"]].dropna().to_string())

# STEP 4 — Free Cash Flow

ko_cf = ko.cashflow.T
ko_cf.index = ko_cf.index.year

fcf = pd.DataFrame()
fcf["Operating Cash Flow"] = (ko_cf["Operating Cash Flow"] / 1_000_000).round(1)
fcf["Capital Expenditure"] = (ko_cf["Capital Expenditure"] / 1_000_000).round(1)
fcf["Free Cash Flow"] = (fcf["Operating Cash Flow"] + fcf["Capital Expenditure"]).round(1)

print("\n=== KO Free Cash Flow ($M) ===")
print(fcf.to_string())

# STEP 5 — KO vs PEP Comparison

print("\n=== KO vs PEP Margin Comparison ===")
print("-- Coca-Cola --")
print(ko_metrics[["Gross Margin %", "Operating Margin %"]].dropna().to_string())
print("\n-- PepsiCo --")
print(pep_metrics[["Gross Margin %", "Operating Margin %"]].dropna().to_string())

# STEP 6 — Export to Excel

wb = Workbook()
ws = wb.active
ws.title = "KO FPandA Report"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="FF0000")

# Section 1 - Income Statement
ws.append(["COCA-COLA FP&A ANALYSIS REPORT"])
ws["A1"].font = Font(bold=True, size=14)
ws.append([])
ws.append(["INCOME STATEMENT ($M)"])
ws.append(["Year", "Total Revenue", "Gross Profit", "Operating Income", "Net Income"])
for cell in ws[ws.max_row]:
    cell.font = header_font
    cell.fill = header_fill
for year, row in ko_metrics[["Total Revenue", "Gross Profit", "Operating Income", "Net Income"]].dropna().iterrows():
    ws.append([year, row["Total Revenue"], row["Gross Profit"], row["Operating Income"], row["Net Income"]])

ws.append([])

# Section 2 - Margins
ws.append(["MARGIN ANALYSIS"])
ws.append(["Year", "Gross Margin %", "Operating Margin %", "Net Margin %", "Revenue Growth %"])
for cell in ws[ws.max_row]:
    cell.font = header_font
    cell.fill = header_fill
for year, row in ko_metrics[["Gross Margin %", "Operating Margin %", "Net Margin %", "Revenue Growth %"]].dropna().iterrows():
    ws.append([year, row["Gross Margin %"], row["Operating Margin %"], row["Net Margin %"], row["Revenue Growth %"]])

ws.append([])

# Section 3 - FCF
ws.append(["FREE CASH FLOW ($M)"])
ws.append(["Year", "Operating Cash Flow", "Capital Expenditure", "Free Cash Flow"])
for cell in ws[ws.max_row]:
    cell.font = header_font
    cell.fill = header_fill
for year, row in fcf.iterrows():
    ws.append([year, row["Operating Cash Flow"], row["Capital Expenditure"], row["Free Cash Flow"]])

ws.append([])

# Section 4 - KO vs PEP
ws.append(["KO vs PEP MARGIN COMPARISON"])
ws.append(["Company", "Year", "Gross Margin %", "Operating Margin %"])
for cell in ws[ws.max_row]:
    cell.font = header_font
    cell.fill = header_fill
for year, row in ko_metrics[["Gross Margin %", "Operating Margin %"]].dropna().iterrows():
    ws.append(["KO", year, row["Gross Margin %"], row["Operating Margin %"]])
for year, row in pep_metrics[["Gross Margin %", "Operating Margin %"]].dropna().iterrows():
    ws.append(["PEP", year, row["Gross Margin %"], row["Operating Margin %"]])

# Auto fit columns
for col in ws.columns:
    max_length = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 4

# Save
output_path = r"C:\Users\chris\Desktop\python\KO_FPandA_Report.xlsx"
wb.save(output_path)

print("\n========================================")
print("  Report exported to KO_FPandA_Report.xlsx")
print("========================================")